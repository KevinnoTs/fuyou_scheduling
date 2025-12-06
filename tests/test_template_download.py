import sys
import os
import unittest
from io import BytesIO
import openpyxl

# Add project root to path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from app import create_app, db
from app.models.users import User
from app.config import Config

class TestConfig(Config):
    TESTING = True
    SQLALCHEMY_DATABASE_URI = 'sqlite:///:memory:'

class TemplateDownloadTestCase(unittest.TestCase):
    def setUp(self):
        # Pass the TestConfig to create_app to ensure we use memory DB
        self.app = create_app(TestConfig)
        self.app.config['WTF_CSRF_ENABLED'] = False
        
        self.client = self.app.test_client()
        
        with self.app.app_context():
            db.create_all()
            # Create admin user
            # Check if exists first to be safe, though drop_all should handle it
            if not User.query.filter_by(username='admin').first():
                u = User(username='admin', real_name='Admin', role='super_admin')
                u.set_password('123456')
                db.session.add(u)
                db.session.commit()
            
    def tearDown(self):
        with self.app.app_context():
            db.session.remove()
            db.drop_all()

    def login(self):
        return self.client.post('/login', data=dict(
            username='admin',
            password='123456'
        ), follow_redirects=True)

    def test_download_template_feb_2025(self):
        self.login()
        # 2025 is not a leap year, Feb has 28 days
        response = self.client.get('/schedule/template?year=2025&month=2')
        self.assertEqual(response.status_code, 200)
        
        wb = openpyxl.load_workbook(BytesIO(response.data))
        ws = wb.active
        
        # Row 1 is header. Name + 1..28 = 29 columns
        self.assertEqual(ws.max_column, 29, "February 2025 should have 29 columns (Name + 28 days)")
        
        # Check last column is '28'
        self.assertEqual(str(ws.cell(row=1, column=29).value), '28')

    def test_download_template_jan_2025(self):
        self.login()
        # Jan has 31 days
        response = self.client.get('/schedule/template?year=2025&month=1')
        self.assertEqual(response.status_code, 200)
        
        wb = openpyxl.load_workbook(BytesIO(response.data))
        ws = wb.active
        
        # Name + 1..31 = 32 columns
        self.assertEqual(ws.max_column, 32, "January 2025 should have 32 columns (Name + 31 days)")
        
        # Check last column is '31'
        self.assertEqual(str(ws.cell(row=1, column=32).value), '31')

if __name__ == '__main__':
    unittest.main()
