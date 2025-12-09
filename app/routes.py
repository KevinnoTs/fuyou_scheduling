import os
import time
import re
from datetime import datetime, date, timedelta
from flask import render_template, flash, redirect, url_for, request, jsonify, send_file
from io import BytesIO
from flask_login import current_user, login_user, logout_user, login_required
from werkzeug.utils import secure_filename
from app import db
from app.models.users import User
from app.models.doctors import Doctor
from app.models.holidays import Holiday
from app.models.schedules import Schedule
from app.models.settings import SystemSetting
import calendar
import openpyxl

def register_routes(app):
    
    @app.before_request
    def check_annual_leave_reset():
        try:
            current_year = datetime.now().year
            # 检查是否已经重置过
            setting = SystemSetting.query.get('last_leave_reset_year')
            
            if not setting or int(setting.value) < current_year:
                # 需要重置
                # 将所有医生的已休年假重置为 0
                doctors = Doctor.query.all()
                for doctor in doctors:
                    doctor.used_annual_leave_days = 0
                
                # 更新设置
                if not setting:
                    setting = SystemSetting(key='last_leave_reset_year', value=str(current_year))
                    db.session.add(setting)
                else:
                    setting.value = str(current_year)
                
                db.session.commit()
                print(f"Annual leave reset for year {current_year}")
        except Exception as e:
            # 防止数据库未初始化时的错误
            print(f"Error in annual leave reset check: {e}")

    
    @app.route('/')
    @app.route('/index')
    def index():
        if current_user.is_authenticated:
            return redirect(url_for('schedule'))
        return render_template('index.html', title='登录')

    @app.route('/login', methods=['POST'])
    def login():
        if current_user.is_authenticated:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': True, 'redirect_url': url_for('schedule')})
            return redirect(url_for('schedule'))
        
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = User.query.filter_by(username=username).first()
        if user is None or not user.check_password(password):
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': '用户名或密码错误'})
            flash('用户名或密码错误')
            return redirect(url_for('index'))
        
        if not user.is_active:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': '该账户已被禁用，请联系管理员'})
            flash('该账户已被禁用，请联系管理员')
            return redirect(url_for('index'))
            
        remember = request.form.get('remember') == 'on'
        login_user(user, remember=remember)

        # Sync theme from cookies if available
        theme_mode = request.cookies.get('theme_mode')
        theme_color = request.cookies.get('theme_color')
        
        if theme_mode or theme_color:
            if theme_mode:
                user.theme_mode = theme_mode
            if theme_color:
                user.theme_color = theme_color
            try:
                db.session.commit()
            except:
                db.session.rollback()

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': True, 'redirect_url': url_for('schedule')})
        return redirect(url_for('schedule'))

    @app.route('/logout')
    def logout():
        logout_user()
        return redirect(url_for('index'))

    @app.route('/schedule')
    @login_required
    def schedule():
        year = request.args.get('year', datetime.now().year, type=int)
        month = request.args.get('month', datetime.now().month, type=int)
        
        # 获取当月有多少天
        _, num_days = calendar.monthrange(year, month)
        
        # 构造日期列表和星期列表
        dates = []
        for day in range(1, num_days + 1):
            date_obj = date(year, month, day)
            dates.append({
                'date': date_obj,
                'day': day,
                'weekday': date_obj.weekday(), # 0=Mon, 6=Sun
                'is_weekend': date_obj.weekday() >= 5,
                'is_today': date_obj == date.today()
            })
            
        # 获取节假日信息
        holidays = Holiday.query.filter(
            db.extract('year', Holiday.date) == year,
            db.extract('month', Holiday.date) == month
        ).all()
        
        holiday_map = {h.date: h for h in holidays}
        
        # 标记日期属性 (节假日/补班)
        for d in dates:
            h = holiday_map.get(d['date'])
            if h:
                d['holiday'] = h
                d['is_workday'] = (h.type == 'workday')
                d['is_holiday'] = (h.type == 'holiday')
            else:
                d['is_workday'] = not d['is_weekend']
                d['is_holiday'] = d['is_weekend']

        # 获取排班数据
        schedules = Schedule.query.filter(
            db.extract('year', Schedule.date) == year,
            db.extract('month', Schedule.date) == month
        ).all()
        
        schedule_map = {} # (doctor_id, day) -> shift_type
        for s in schedules:
            schedule_map[(s.doctor_id, s.date.day)] = s.shift_type
            
        # 获取医生并分组
        all_doctors = Doctor.query.order_by(Doctor.display_order.asc()).all()
        doctors_group = []
        assistants_group = []
        
        for doc in all_doctors:
            # 过滤逻辑：如果医生已离职，且浏览月份晚于离职月份，则不显示
            if doc.status == 'resigned' and doc.resignation_date:
                # 构造浏览月份的第一天
                view_date = date(year, month, 1)
                # 构造离职月份的第一天 (忽略具体的日，只要月份对齐)
                resign_month_date = date(doc.resignation_date.year, doc.resignation_date.month, 1)
                
                if view_date > resign_month_date:
                    continue

            # 简单的分组逻辑：职位含"医助"分为一组，其他为医生组
            # 或者根据 display_order? 这里暂时用 Title
            if doc.title and '医助' in doc.title:
                assistants_group.append(doc)
            else:
                doctors_group.append(doc)
                
        return render_template('schedule.html', 
                             title='排班表',
                             year=year,
                             month=month,
                             dates=dates,
                             doctors_group=doctors_group,
                             assistants_group=assistants_group,
                             schedule_map=schedule_map)

    @app.route('/schedule/upload', methods=['POST'])
    @login_required
    def upload_schedule():
        if current_user.role not in ['admin', 'super_admin']:
             return jsonify({'success': False, 'message': '权限不足'})

        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '未上传文件'})
            
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': '未选择文件'})
            
        if file.filename.endswith('.xls'):
            return jsonify({'success': False, 'message': '不支持旧版 Excel (.xls) 格式，请另存为 .xlsx 格式后上传。'})

        if not file.filename.endswith('.xlsx'):
            return jsonify({'success': False, 'message': '仅支持 Excel 2007+ (.xlsx) 文件'})
            
        temp_path = None
        try:
            # 1. 保存到临时文件
            filename = secure_filename(file.filename)
            temp_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'temp')
            if not os.path.exists(temp_dir):
                os.makedirs(temp_dir)
            temp_path = os.path.join(temp_dir, f"{int(time.time())}_{filename}")
            file.save(temp_path)

            # 2. 读取文件 (data_only=True 读取值而非公式)
            try:
                wb = openpyxl.load_workbook(temp_path, data_only=True)
            except Exception as load_err:
                 return jsonify({'success': False, 'message': f'文件读取失败: 可能是文件加密或损坏 ({str(load_err)})'})

            sheet = wb.active
            
            # 3. 获取前端传递的参数
            year = request.form.get('year', datetime.now().year, type=int)
            month = request.form.get('month', datetime.now().month, type=int)
            
            # 读取所有医生做映射 Name -> ID
            doctors = Doctor.query.all()
            doctor_map = {d.name: d.id for d in doctors}
            
            updated_count = 0
            
            # 遍历行
            for row in sheet.iter_rows(min_row=2, values_only=True): 
                # row[0] 是姓名 (假设格式)
                if not row or len(row) < 2:
                    continue

                name = row[0]
                if not name or name not in doctor_map:
                    continue
                    
                doctor_id = doctor_map[name]
                
                # 假设第一行是日期，列索引1对应1号 (Excel中第2列)
                # row[1] -> 1号, row[2] -> 2号 ...
                
                for day_idx in range(1, len(row)):
                    day = day_idx 
                    try:
                        # 检查日期有效性
                        current_date = date(year, month, day)
                    except ValueError:
                        break # 超出当月天数
                        
                    shift_type = row[day_idx]
                    if shift_type:
                        shift_type = str(shift_type).strip()
                    
                    # 查找现有记录
                    schedule = Schedule.query.filter_by(doctor_id=doctor_id, date=current_date).first()
                    
                    if shift_type:
                        if schedule:
                            schedule.shift_type = shift_type
                        else:
                            schedule = Schedule(doctor_id=doctor_id, date=current_date, shift_type=shift_type)
                            db.session.add(schedule)
                        updated_count += 1
                    elif schedule:
                         # 如果 Excel 为空但数据库有值，暂不处理（或根据需求清空）
                         pass
            
            db.session.commit()
            return jsonify({'success': True, 'message': f'导入成功，更新了 {updated_count} 条排班信息'})
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'message': f'系统错误: {str(e)}'})
        finally:
            # 清理临时文件
            if temp_path and os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except:
                    pass

    @app.route('/schedule/doctor/<int:doctor_id>')
    @login_required
    def doctor_schedule_calendar(doctor_id):
        year = request.args.get('year', datetime.now().year, type=int)
        month = request.args.get('month', datetime.now().month, type=int)
        
        doctor = db.session.get(Doctor, doctor_id)
        if not doctor:
            flash('医生不存在')
            return redirect(url_for('schedule'))
            
        _, num_days = calendar.monthrange(year, month)
        
        # Get Holidays
        holidays = Holiday.query.filter(
            db.extract('year', Holiday.date) == year,
            db.extract('month', Holiday.date) == month
        ).all()
        holiday_map = {h.date.day: h for h in holidays}
        
        # Get Schedules
        schedules = Schedule.query.filter(
            Schedule.doctor_id == doctor_id,
            db.extract('year', Schedule.date) == year,
            db.extract('month', Schedule.date) == month
        ).all()
        schedule_map = {s.date.day: s.shift_type for s in schedules}
        
        # Build Calendar Data
        calendar_data = []
        start_day_weekday = date(year, month, 1).weekday() # 0=Mon, 6=Sun
        
        # Empty cells for start of month
        for _ in range(start_day_weekday):
            calendar_data.append(None)
            
        for day in range(1, num_days + 1):
            current_date = date(year, month, day)
            is_weekend = current_date.weekday() >= 5
            
            status = 'weekend' if is_weekend else 'workday'
            holiday_name = None
            
            if day in holiday_map:
                h = holiday_map[day]
                if h.type == 'holiday':
                    status = 'holiday'
                    holiday_name = h.name
                elif h.type == 'workday':
                    status = 'workday'
                    holiday_name = h.name
            
            calendar_data.append({
                'day': day,
                'status': status,
                'holiday_name': holiday_name,
                'shift_type': schedule_map.get(day)
            })
            
        # Fill end of week
        while len(calendar_data) % 7 != 0:
            calendar_data.append(None)
            
        return render_template('doctor_schedule_calendar.html',
                             doctor=doctor,
                             current_year=year,
                             current_month=month,
                             calendar_data=calendar_data)

    @app.route('/schedule/template')
    @login_required
    def download_schedule_template():
        if current_user.role not in ['admin', 'super_admin']:
             flash('权限不足')
             return redirect(url_for('schedule'))
        
        # Get year/month from query params, default to current
        try:
            year = int(request.args.get('year', datetime.now().year))
            month = int(request.args.get('month', datetime.now().month))
        except ValueError:
            year = datetime.now().year
            month = datetime.now().month

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{year}年{month}月排班模版"
        
        # 1. Calculate number of days
        _, num_days = calendar.monthrange(year, month)

        # 2. Key Headers
        headers = ['姓名'] + [str(i) for i in range(1, num_days + 1)]
        ws.append(headers)
        
        # 3. Pre-fill Doctor Names
        doctors = Doctor.query.order_by(Doctor.display_order.asc()).all()
        for doc in doctors:
            if doc.status == 'active':
                # Create row with Name + empty cells for each day
                row = [doc.name] + [''] * num_days
                ws.append(row)
                
        # 4. Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'schedule_template_{year}_{month:02d}.xlsx'

        return send_file(
            output, 
            as_attachment=True, 
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    @app.route('/doctors')
    @login_required
    def doctors_list():
        # Always return all doctors for client-side filtering
        doctors = Doctor.query.order_by(Doctor.display_order.asc()).all()
        return render_template('doctors_list.html', title='医生列表', doctors=doctors)

    @app.route('/users')
    @login_required
    def users_list():
        if current_user.role not in ['admin', 'super_admin']:
            flash('您没有权限访问该页面')
            return redirect(url_for('index'))
        
        search_query = request.args.get('q', '')
        if search_query:
            users = User.query.filter(
                (User.username.contains(search_query)) | 
                (User.real_name.contains(search_query))
            ).all()
        else:
            users = User.query.all()
            
        return render_template('users_list.html', title='用户列表', users=users)

    @app.route('/users/add', methods=['GET', 'POST'])
    @login_required
    def add_user():
        if current_user.role not in ['admin', 'super_admin']:
            flash('您没有权限执行此操作')
            return redirect(url_for('users_list'))
            
        if request.method == 'POST':
            username = request.form.get('username')
            real_name = request.form.get('real_name')
            role = request.form.get('role')
            password = request.form.get('password')
            
            # 权限检查：普通管理员只能创建普通用户或普通医助
            if current_user.role == 'admin' and role not in ['user', 'assistant']:
                flash('管理员只能创建普通医师或普通医助')
                doctors = Doctor.query.order_by(Doctor.display_order.asc()).all()
                return render_template('user_form.html', title='添加用户', doctors=doctors)

            if User.query.filter_by(username=username).first():
                flash('用户名已存在')
                doctors = Doctor.query.order_by(Doctor.display_order.asc()).all()
                return render_template('user_form.html', title='添加用户', doctors=doctors)
            
            user = User(username=username, real_name=real_name, role=role)
            
            # 处理关联医生
            doctor_id = request.form.get('doctor_id')
            if doctor_id:
                user.doctor_id = int(doctor_id)
                
            user.set_password(password)
            db.session.add(user)
            db.session.commit()
            
            flash('用户添加成功')
            return redirect(url_for('users_list'))
            
        doctors = Doctor.query.order_by(Doctor.display_order.asc()).all()
        return render_template('user_form.html', title='添加用户', doctors=doctors)

    @app.route('/users/edit/<int:id>', methods=['GET', 'POST'])
    @login_required
    def edit_user(id):
        if current_user.role not in ['admin', 'super_admin']:
            flash('您没有权限执行此操作')
            return redirect(url_for('users_list'))
            
        user = db.session.get(User, id)
        if not user:
            flash('用户不存在')
            return redirect(url_for('users_list'))
            
        # 权限检查：管理员只能编辑普通用户
        # 权限检查：管理员只能编辑普通用户或普通医助，或者自己
        if current_user.role == 'admin' and user.role not in ['user', 'assistant'] and user.id != current_user.id:
            flash('您没有权限编辑此账户')
            return redirect(url_for('users_list'))

        if request.method == 'POST':
            real_name = request.form.get('real_name')
            role = request.form.get('role')
            password = request.form.get('password')
            is_active = request.form.get('is_active') == 'on'
            
            # 防止普通管理员将自己或他人提升为超级管理员
            if role == 'super_admin' and current_user.role != 'super_admin':
                flash('无法设置为超级管理员')
                return render_template('user_form.html', title='编辑用户', user=user)

            user.real_name = real_name
            user.is_active_user = is_active
            
            # 只有超级管理员可以修改角色
            if current_user.role == 'super_admin':
                user.role = role
            
            # 如果提供了密码，则修改密码
            if password and password.strip():
                user.set_password(password)
                flash('密码已修改')
            
            # 关联医生 (仅管理员可操作)
            if current_user.role in ['admin', 'super_admin']:
                doctor_id = request.form.get('doctor_id')
                if doctor_id:
                    user.doctor_id = int(doctor_id)
                else:
                    user.doctor_id = None

            db.session.commit()
            flash('用户信息已更新')
            return redirect(url_for('users_list'))
            
        doctors = Doctor.query.order_by(Doctor.display_order.asc()).all()
        return render_template('user_form.html', title='编辑用户', user=user, doctors=doctors)

    @app.route('/users/delete/<int:id>', methods=['POST'])
    @login_required
    def delete_user(id):
        if current_user.role not in ['admin', 'super_admin']:
            flash('您没有权限执行此操作')
            return redirect(url_for('users_list'))
            
        user = db.session.get(User, id)
        if user:
            # 权限检查
            if current_user.role == 'admin' and user.role not in ['user', 'assistant']:
                flash('权限不足')
                return redirect(url_for('users_list'))

            if user.role == 'super_admin':
                flash('无法删除超级管理员')
            elif user.id == current_user.id:
                flash('无法删除自己')
            else:
                db.session.delete(user)
                db.session.commit()
                flash('用户已删除')
        else:
            flash('用户不存在')
        return redirect(url_for('users_list'))

    @app.route('/doctors/add', methods=['GET', 'POST'])
    @login_required
    def add_doctor():
        if current_user.role not in ['admin', 'super_admin']:
            flash('您没有权限执行此操作')
            return redirect(url_for('doctors_list'))
        
        if request.method == 'POST':
            name = request.form.get('name')
            gender = request.form.get('gender')
            title = request.form.get('title')
            # 处理多选擅长方向
            specialties = request.form.getlist('specialty')
            specialty = ','.join(specialties) if specialties else ''
            
            status = request.form.get('status')
            annual_leave_days = int(request.form.get('annual_leave_days', 5))
            used_annual_leave_days = int(request.form.get('used_annual_leave_days', 0))
            display_order = int(request.form.get('display_order', 999))
            
            # 验证年假逻辑
            if used_annual_leave_days > annual_leave_days:
                flash('错误：已休年假天数不能大于每年总年假天数')
                # 保留用户输入的数据以便回显
                doctor = Doctor(
                    name=name,
                    gender=gender,
                    title=title,
                    specialty=specialty,
                    status=status,
                    annual_leave_days=annual_leave_days,
                    used_annual_leave_days=used_annual_leave_days,
                    display_order=display_order
                )
                return render_template('doctor_form.html', title='添加医生', doctor=doctor)

            # 处理头像上传
            avatar_path = None
            if 'avatar' in request.files:
                file = request.files['avatar']
                if file and file.filename != '' and file.filename.split('.')[-1].lower() in app.config['ALLOWED_EXTENSIONS']:

                    filename = secure_filename(file.filename)
                    # 使用时间戳防止重名

                    filename = f"{int(time.time())}_{filename}"
                    file.save(os.path.join(app.config['UPLOAD_FOLDER'], 'avatars', filename))
                    avatar_path = f"uploads/avatars/{filename}"

            doctor = Doctor(
                name=name,
                gender=gender,
                title=title,
                specialty=specialty,
                status=status,
                annual_leave_days=annual_leave_days,
                used_annual_leave_days=used_annual_leave_days,
                display_order=display_order,
                avatar_path=avatar_path
            )
            
            if status == 'resigned':
                try:
                    r_year = int(request.form.get('resignation_year'))
                    r_month = int(request.form.get('resignation_month'))
                    doctor.resignation_date = date(r_year, r_month, 1)
                except (ValueError, TypeError):
                    pass
            
            db.session.add(doctor)
            db.session.commit()
            flash('医生添加成功')
            return redirect(url_for('doctors_list'))
            
        return render_template('doctor_form.html', title='添加医生')

    @app.route('/doctors/edit/<int:id>', methods=['GET', 'POST'])
    @login_required
    def edit_doctor(id):
        if current_user.role not in ['admin', 'super_admin']:
            flash('您没有权限执行此操作')
            return redirect(url_for('doctors_list'))
            
        doctor = db.session.get(Doctor, id)
        if not doctor:
            flash('医生不存在')
            return redirect(url_for('doctors_list'))
            
        if request.method == 'POST':
            # 获取表单数据
            name = request.form.get('name')
            gender = request.form.get('gender')
            title = request.form.get('title')
            specialties = request.form.getlist('specialty')
            specialty = ','.join(specialties) if specialties else ''
            status = request.form.get('status')
            annual_leave_days = int(request.form.get('annual_leave_days', 5))
            used_annual_leave_days = int(request.form.get('used_annual_leave_days', 0))
            display_order = int(request.form.get('display_order', 999))

            # 验证年假逻辑
            if used_annual_leave_days > annual_leave_days:
                flash('错误：已休年假天数不能大于每年总年假天数')
                # 更新对象属性以便回显，但不提交到数据库
                doctor.name = name
                doctor.gender = gender
                doctor.title = title
                doctor.specialty = specialty
                doctor.status = status
                doctor.annual_leave_days = annual_leave_days
                doctor.used_annual_leave_days = used_annual_leave_days
                doctor.display_order = display_order
                return render_template('doctor_form.html', title='编辑医生', doctor=doctor)

            doctor.name = name
            doctor.gender = gender
            doctor.title = title
            doctor.specialty = specialty
            doctor.status = status
            doctor.annual_leave_days = annual_leave_days
            doctor.used_annual_leave_days = used_annual_leave_days
            doctor.display_order = display_order
            
            # 处理头像上传
            if 'avatar' in request.files:
                file = request.files['avatar']
                if file and file.filename != '' and file.filename.split('.')[-1].lower() in app.config['ALLOWED_EXTENSIONS']:

                    filename = secure_filename(file.filename)

                    filename = f"{int(time.time())}_{filename}"
                    file.save(os.path.join(app.config['UPLOAD_FOLDER'], 'avatars', filename))
                    doctor.avatar_path = f"uploads/avatars/{filename}"

            if doctor.status == 'resigned':
                try:
                    r_year = int(request.form.get('resignation_year'))
                    r_month = int(request.form.get('resignation_month'))
                    doctor.resignation_date = date(r_year, r_month, 1)
                except (ValueError, TypeError):
                    pass
            else:
                 # 如果改回在职，重置离职时间为默认

                 doctor.resignation_date = date(2099, 12, 31)

            db.session.commit()
            flash('医生信息已更新')
            return redirect(url_for('doctors_list'))
            
        return render_template('doctor_form.html', title='编辑医生', doctor=doctor)

    @app.route('/doctors/delete/<int:id>', methods=['POST'])
    @login_required
    def delete_doctor(id):
        if current_user.role not in ['admin', 'super_admin']:
            flash('您没有权限执行此操作')
            return redirect(url_for('doctors_list'))
            
        doctor = db.session.get(Doctor, id)
        if doctor:
            db.session.delete(doctor)
            db.session.commit()
            flash('医生已删除')
        else:
            flash('医生不存在')
        return redirect(url_for('doctors_list'))
        return redirect(url_for('doctors_list'))

    @app.route('/doctors/upload_avatar/<int:id>', methods=['POST'])
    @login_required
    def upload_doctor_avatar(id):
        doctor = db.session.get(Doctor, id)
        if not doctor:
            flash('医生不存在')
            return redirect(url_for('doctors_list'))
            
        # 权限检查：只有管理员或关联的用户可以修改头像
        if current_user.role not in ['admin', 'super_admin'] and current_user.doctor_id != doctor.id:
            flash('您没有权限修改此医生的头像')
            return redirect(url_for('doctors_list'))
            
        if 'avatar' in request.files:
            file = request.files['avatar']
            if file and file.filename != '' and file.filename.split('.')[-1].lower() in app.config['ALLOWED_EXTENSIONS']:

                filename = secure_filename(file.filename)

                filename = f"{int(time.time())}_{filename}"
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], 'avatars', filename))
                doctor.avatar_path = f"uploads/avatars/{filename}"
                db.session.commit()
                flash('头像已更新')
            else:
                flash('上传失败：无效的文件或格式')
        
        return redirect(url_for('doctors_list'))

    # -------------------------------------------------------------------------
    # 节假日管理模块
    # -------------------------------------------------------------------------
    @app.route('/holidays')
    @login_required
    def holidays_list():
        year = request.args.get('year', datetime.now().year, type=int)
        holidays = Holiday.query.filter(
            db.extract('year', Holiday.date) == year
        ).order_by(Holiday.date.asc()).all()
        
        # Calculate total records count
        total_count = len(holidays)
        
        return render_template('holidays_list.html', 
                             title='节假日管理', 
                             holidays=holidays, 
                             current_year=year,
                             total_count=total_count)

    @app.route('/holidays/add', methods=['GET', 'POST'])
    @login_required
    def add_holiday():
        if current_user.role not in ['admin', 'super_admin']:
            flash('您没有权限执行此操作')
            return redirect(url_for('holidays_list'))
            
        if request.method == 'POST':
            start_date_str = request.form.get('start_date')
            end_date_str = request.form.get('end_date')
            name = request.form.get('name')
            type = request.form.get('type')
            
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = start_date # Default to single day
                
                if end_date_str:
                    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                
                if end_date < start_date:
                    flash('结束日期不能早于开始日期')
                    return render_template('holiday_form.html', title='添加节假日')

                # Iterate and add
                curr = start_date

                added_count = 0
                
                while curr <= end_date:
                    if not Holiday.query.filter_by(date=curr).first():
                        holiday = Holiday(date=curr, name=name, type=type, source='manual')
                        db.session.add(holiday)
                        added_count += 1
                    curr += timedelta(days=1)
                
                db.session.commit()
                flash(f'成功添加 {added_count} 天节假日')
                return redirect(url_for('holidays_list', year=start_date.year))
                
            except ValueError:
                flash('无效的日期格式')
                
        return render_template('holiday_form.html', title='添加节假日')

    @app.route('/holidays/edit/<int:id>', methods=['GET', 'POST'])
    @login_required
    def edit_holiday(id):
        if current_user.role not in ['admin', 'super_admin']:
            flash('您没有权限执行此操作')
            return redirect(url_for('holidays_list'))
            
        holiday = db.session.get(Holiday, id)
        if not holiday:
            flash('节假日不存在')
            return redirect(url_for('holidays_list'))
            
        if request.method == 'POST':
            name = request.form.get('name')
            type = request.form.get('type')
            # Date is usually not editable to avoid conflicts, but if needed we can add it.
            # For now, let's allow editing name and type.
            
            holiday.name = name
            holiday.type = type
            db.session.commit()
            flash('节假日已更新')
            return redirect(url_for('holidays_list', year=holiday.date.year))
            
        return render_template('holiday_form.html', title='编辑节假日', holiday=holiday)

    @app.route('/holidays/delete/<int:id>', methods=['POST'])
    @login_required
    def delete_holiday(id):
        if current_user.role not in ['admin', 'super_admin']:
            flash('您没有权限执行此操作')
            return redirect(url_for('holidays_list'))
            
        holiday = db.session.get(Holiday, id)
        if holiday:
            year = holiday.date.year
            db.session.delete(holiday)
            db.session.commit()
            flash('节假日已删除')
            return redirect(url_for('holidays_list', year=year))
        
        flash('节假日不存在')
        return redirect(url_for('holidays_list'))

    @app.route('/holidays/import_text', methods=['POST'])
    @login_required
    def import_holidays_text():
        if current_user.role not in ['admin', 'super_admin']:
            return jsonify({'success': False, 'message': '权限不足'})
            
        text = request.form.get('text', '')
        # year param is optional now, we try to detect it
        year_param = request.form.get('year', type=int)
        
        if not text:
            return jsonify({'success': False, 'message': '缺少文本内容'})
            
        try:

            
            # 1. Detect Year
            year_match = re.search(r'(\d{4})年', text)
            if year_match:
                year = int(year_match.group(1))
            elif year_param:
                year = year_param
            else:
                return jsonify({'success': False, 'message': '无法识别年份，请确保文本包含年份（如“2025年”）'})

            # Normalize newlines
            text = text.replace('\r\n', '\n').replace('\r', '\n')
            
            # Split by segments using a more robust method
            # We insert a special separator before "Number + 、" at the start of lines or after whitespace
            processed_text = re.sub(r'(?:^|\s)([一二三四五六七八九十]+、)', r'__SEGMENT__\1', text)
            segments = processed_text.split('__SEGMENT__')
            
            added_count = 0
            
            # Regex for dates
            date_pattern = re.compile(r'(\d{1,2})月(\d{1,2})日')
            
            for segment in segments:
                if not segment.strip():
                    continue
                
                # Remove the numbering "一、"
                segment = re.sub(r'^[一二三四五六七八九十]+、', '', segment.strip())
                
                # Remove newlines within segment to ensure regex works across lines
                segment = segment.replace('\n', '').strip()
                
                # Preprocess text within segment
                segment = segment.replace('：', ':').replace('，', ',').replace('。', '.').replace('、', ',')
                
                if ':' not in segment:
                    continue
                    
                # Extract Name
                name_part, content_part = segment.split(':', 1)
                holiday_name = name_part.strip()
                
                # Separate "Holiday" and "Workday" parts
                work_keyword_index = content_part.find('上班')
                
                holiday_part = content_part
                workday_part = ""
                
                if work_keyword_index != -1:
                    parts = re.split(r'[.;]', content_part)
                    h_parts = []
                    w_parts = []
                    for p in parts:
                        if '上班' in p:
                            w_parts.append(p)
                        else:
                            h_parts.append(p)
                    holiday_part = " ".join(h_parts)
                    workday_part = " ".join(w_parts)
                
                # --- Parse Holiday Dates ---
                
                # 1. Range with Month change: "1月28日...至2月4日"
                ranges_full = re.findall(r'(\d{1,2})月(\d{1,2})日(?:(?!\d{1,2}月\d{1,2}日).)*?至(?:(?!\d{1,2}月\d{1,2}日).)*?(\d{1,2})月(\d{1,2})日', holiday_part, re.DOTALL)
                for m1, d1, m2, d2 in ranges_full:
                    start = date(year, int(m1), int(d1))
                    end = date(year, int(m2), int(d2))
                    curr = start
                    while curr <= end:
                        if not Holiday.query.filter_by(date=curr).first():
                            db.session.add(Holiday(date=curr, name=holiday_name, type='holiday', source='manual'))
                            added_count += 1
                        curr += timedelta(days=1)

                # 2. Range within same month: "4月4日...至6日"
                # Strict regex to avoid matching across months or other numbers
                ranges_short_strict = re.findall(r'(\d{1,2})月(\d{1,2})日(?:(?!\d{1,2}月\d{1,2}日).)*?至(?:(?!\d{1,2}月\d{1,2}日).)*?(\d{1,2})日', holiday_part, re.DOTALL)
                for m, d1, d2 in ranges_short_strict:
                    start = date(year, int(m), int(d1))
                    end = date(year, int(m), int(d2))
                    curr = start
                    while curr <= end:
                        if not Holiday.query.filter_by(date=curr).first():
                            db.session.add(Holiday(date=curr, name=holiday_name, type='holiday', source='manual'))
                            added_count += 1
                        curr += timedelta(days=1)

                # 3. Single Days
                singles = date_pattern.findall(holiday_part)
                for m, d in singles:
                    curr = date(year, int(m), int(d))
                    # Only add if not exists (and thus not added by range logic above)
                    if not Holiday.query.filter_by(date=curr).first():
                        db.session.add(Holiday(date=curr, name=holiday_name, type='holiday', source='manual'))
                        added_count += 1

                # --- Parse Workdays ---
                if workday_part:
                    workdays = date_pattern.findall(workday_part)
                    for m, d in workdays:
                        curr = date(year, int(m), int(d))
                        # Check if exists
                        existing = Holiday.query.filter_by(date=curr).first()
                        if existing:
                            if existing.type == 'holiday':
                                existing.type = 'workday'
                                existing.name = f"{holiday_name}调休"
                        else:
                            db.session.add(Holiday(date=curr, name=f"{holiday_name}调休", type='workday', source='manual'))
                            added_count += 1
                            
            db.session.commit()
            return jsonify({'success': True, 'message': f'成功导入 {added_count} 条数据 (年份: {year})'})
            
        except Exception as e:
            return jsonify({'success': False, 'message': str(e)})

    @app.route('/api/update_theme', methods=['POST'])
    @login_required
    def update_theme():
        data = request.get_json()
        mode = data.get('mode')
        color = data.get('color')

        if mode:
            current_user.theme_mode = mode
        if color:
            current_user.theme_color = color
        
        try:
            db.session.commit()
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500
