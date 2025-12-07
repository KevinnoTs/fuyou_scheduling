from flask import render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from app import db
from app.models.doctors import Doctor
from app.models.workloads import WorkloadScore, WorkloadRecord
from datetime import datetime, date, timedelta
from sqlalchemy import func, or_, and_
import openpyxl
import xlrd
from io import BytesIO
from urllib.parse import quote
from pypinyin import lazy_pinyin

def register_workload_routes(app):
    
    def get_date_params():
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        
        if not year or not month:
            # Try to find the latest date with data
            latest_date = db.session.query(func.max(WorkloadRecord.report_date)).scalar()
            
            if latest_date:
                # If data exists, use the latest year/month
                year = latest_date.year
                month = latest_date.month
            else:
                # If no data, default to current month
                current_date = datetime.now()
                year = current_date.year
                month = current_date.month
            
        return year, month

    @app.route('/workload')
    @login_required
    def workload_stats():
        if current_user.role not in ['doctor', 'admin', 'super_admin']:
            flash('无权访问工作量统计页面')
            return redirect(url_for('index'))
            
        import calendar
        year, month = get_date_params()
        month_str = f"{year}-{month:02d}"
        selected_month_start = date(year, month, 1)

        # Expansion Logic
        expand_mode = request.args.get('expand') == '1'
        days_in_month = 0
        daily_scores_map = {} # doctor_id -> {day: score}

        if expand_mode:
            _, days_in_month = calendar.monthrange(year, month)
            # Fetch all records with scores for the month
            records = db.session.query(WorkloadRecord.doctor_id, WorkloadRecord.report_date, WorkloadScore.score)\
                .join(WorkloadScore, WorkloadRecord.item_name == WorkloadScore.item_name)\
                .filter(func.strftime('%Y-%m', WorkloadRecord.report_date) == month_str)\
                .all()
            
            for doc_id, r_date, score in records:
                if not doc_id or not score: continue
                day = r_date.day
                if doc_id not in daily_scores_map:
                    daily_scores_map[doc_id] = {}
                daily_scores_map[doc_id][day] = daily_scores_map[doc_id].get(day, 0) + score

        # Get doctors
        doctors = Doctor.query.filter(
            and_(
                Doctor.title == '医师',
                or_(
                    Doctor.status == 'active',
                    and_(
                        Doctor.status != 'active',
                        Doctor.resignation_date >= selected_month_start
                    )
                )
            )
        ).order_by(Doctor.display_order.asc()).all()
        
        stats = []
        for doctor in doctors:
            total_score = db.session.query(func.sum(WorkloadScore.score))\
                .select_from(WorkloadRecord)\
                .join(WorkloadScore, WorkloadRecord.item_name == WorkloadScore.item_name)\
                .filter(WorkloadRecord.doctor_id == doctor.id)\
                .filter(func.strftime('%Y-%m', WorkloadRecord.report_date) == month_str)\
                .scalar() or 0
            
            item = {
                'doctor': doctor,
                'total_score': round(total_score, 2)
            }
            if expand_mode:
                # Prepare a list or dict for template. Dict is easier for lookups.
                doc_daily = daily_scores_map.get(doctor.id, {})
                # Round daily scores
                item['daily_scores'] = {d: round(s, 2) for d, s in doc_daily.items()}
            
            stats.append(item)
            
        return render_template('workload_stats.html', 
                             stats=stats, 
                             current_year=year,
                             current_month=month,
                             expand_mode=expand_mode,
                             days_in_month=days_in_month,
                             title='工作量统计')

    @app.route('/workload/records', methods=['GET', 'POST'])
    @login_required
    def workload_records():
        year, month = get_date_params()
        month_str = f"{year}-{month:02d}"
        
        if request.method == 'POST':
            if current_user.role not in ['admin', 'super_admin']:
                flash('无权操作')
                return redirect(url_for('workload_records'))
                
            # Manual Add Logic
            doctor_name = request.form.get('doctor_name')
            item_name = request.form.get('item_name')
            report_date_str = request.form.get('report_date')
            
            if doctor_name and item_name and report_date_str:
                try:
                    report_date = datetime.strptime(report_date_str, '%Y-%m-%d').date()
                    # Try to find doctor by name to link ID
                    doctor = Doctor.query.filter_by(name=doctor_name).first()
                    
                    record = WorkloadRecord(
                        doctor_name=doctor_name,
                        doctor_id=doctor.id if doctor else None,
                        item_name=item_name,
                        report_date=report_date
                    )
                    db.session.add(record)
                    db.session.commit()
                    flash('记录添加成功')
                except Exception as e:
                    flash(f'添加失败: {str(e)}')
            
            return redirect(url_for('workload_records', year=year, month=month))

        # List records
        query = db.session.query(WorkloadRecord, WorkloadScore.score)\
            .outerjoin(WorkloadScore, WorkloadRecord.item_name == WorkloadScore.item_name)\
            .outerjoin(Doctor, WorkloadRecord.doctor_id == Doctor.id)\
            .filter(func.strftime('%Y-%m', WorkloadRecord.report_date) == month_str)\
            .order_by(WorkloadRecord.report_date.asc(), Doctor.display_order.asc())
            
        records = query.all()
        
        return render_template('workload_records.html', 
                             records=records, 
                             current_year=year,
                             current_month=month,
                             title='工作记录')

    @app.route('/workload/import_records', methods=['POST'])
    @login_required
    def import_workload_records():
        if current_user.role not in ['admin', 'super_admin']:
            flash('无权操作')
            return redirect(url_for('workload_records'))
            
        if 'file' not in request.files:
            flash('未上传文件')
            return redirect(url_for('workload_records'))
            
        file = request.files['file']
        if file.filename == '':
            flash('未选择文件')
            return redirect(url_for('workload_records'))
            
        if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
            try:
                rows = []
                
                if file.filename.endswith('.xlsx'):
                    wb = openpyxl.load_workbook(file)
                    ws = wb.active
                    # Get all rows as list of values
                    rows = list(ws.iter_rows(values_only=True))
                else:
                    # Handle .xls with xlrd
                    file_content = file.read()
                    wb = xlrd.open_workbook(file_contents=file_content)
                    sheet = wb.sheet_by_index(0)
                    rows = [sheet.row_values(i) for i in range(sheet.nrows)]
                
                if not rows:
                    flash('文件为空')
                    return redirect(url_for('workload_records'))

                # Header mapping
                header_row = rows[0]
                
                required_cols = {
                    'report_date': ['报告日期'],
                    'doctor_name': ['报告医师'],
                    'item_name': ['检查项目']
                }
                
                col_indices = {}
                
                for idx, col_name in enumerate(header_row):
                    if not col_name: continue
                    col_name = str(col_name).strip()
                    for key, aliases in required_cols.items():
                        if col_name in aliases:
                            col_indices[key] = idx
                            break
                            
                # Check missing columns
                missing = [k for k in required_cols if k not in col_indices]
                if missing:
                    flash(f'缺少必要列: {", ".join(missing)} (需包含: 报告日期, 报告医师, 检查项目)')
                    return redirect(url_for('workload_records'))
                    
                success_count = 0
                skip_count = 0
                
                for row in rows[1:]:
                    try:
                        # Extract data
                        # Handle case where row might be shorter than header
                        if len(row) <= max(col_indices.values()):
                            skip_count += 1
                            continue
                            
                        r_date_val = row[col_indices['report_date']]
                        doc_name = row[col_indices['doctor_name']]
                        item_name = row[col_indices['item_name']]
                        
                        if not r_date_val or not doc_name or not item_name:
                            skip_count += 1
                            continue
                            
                        # Parse Date
                        report_date = None
                        if isinstance(r_date_val, datetime):
                            report_date = r_date_val.date()
                        elif isinstance(r_date_val, date):
                            report_date = r_date_val
                        elif isinstance(r_date_val, float):
                            # Handle Excel serial date from xlrd (float)
                            try:
                                dt_tuple = xlrd.xldate_as_tuple(r_date_val, wb.datemode if hasattr(wb, 'datemode') else 0)
                                report_date = date(dt_tuple[0], dt_tuple[1], dt_tuple[2])
                            except:
                                pass
                        elif isinstance(r_date_val, str):
                            # Try common formats
                            for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y%m%d']:
                                try:
                                    report_date = datetime.strptime(r_date_val, fmt).date()
                                    break
                                except ValueError:
                                    continue
                        
                        if not report_date:
                            skip_count += 1
                            continue
                            
                        # Find Doctor ID
                        doctor = Doctor.query.filter_by(name=doc_name).first()
                        
                        record = WorkloadRecord(
                            doctor_name=doc_name,
                            doctor_id=doctor.id if doctor else None,
                            item_name=item_name,
                            report_date=report_date
                        )
                        db.session.add(record)
                        success_count += 1
                        
                    except Exception as e:
                        skip_count += 1
                        continue
                        
                db.session.commit()
                flash(f'导入成功 {success_count} 条，跳过 {skip_count} 条')
                
            except Exception as e:
                flash(f'导入失败: {str(e)}')
        else:
            flash('不支持的文件格式')
            
        return redirect(url_for('workload_records'))

    @app.route('/workload/import_scores', methods=['POST'])
    @login_required
    def import_workload_scores():
        if current_user.role not in ['admin', 'super_admin']:
            flash('无权操作')
            return redirect(url_for('workload_settings'))
            
        if 'file' not in request.files:
            flash('未上传文件')
            return redirect(url_for('workload_settings'))
            
        file = request.files['file']
        if file.filename == '':
            flash('未选择文件')
            return redirect(url_for('workload_settings'))
            
        if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
            try:
                rows = []
                
                if file.filename.endswith('.xlsx'):
                    wb = openpyxl.load_workbook(file)
                    ws = wb.active
                    rows = list(ws.iter_rows(values_only=True))
                else:
                    file_content = file.read()
                    wb = xlrd.open_workbook(file_contents=file_content)
                    sheet = wb.sheet_by_index(0)
                    rows = [sheet.row_values(i) for i in range(sheet.nrows)]
                
                if not rows:
                    flash('文件为空')
                    return redirect(url_for('workload_settings'))

                # Header mapping
                header_row = rows[0]
                
                required_cols = {
                    'item_name': ['检查项目', '项目名称'],
                    'score': ['分值', '分数']
                }
                
                col_indices = {}
                
                for idx, col_name in enumerate(header_row):
                    if not col_name: continue
                    col_name = str(col_name).strip()
                    for key, aliases in required_cols.items():
                        if col_name in aliases:
                            col_indices[key] = idx
                            break
                            
                missing = [k for k in required_cols if k not in col_indices]
                if missing:
                    flash(f'缺少必要列: {", ".join(missing)} (需包含: 检查项目, 分值)')
                    return redirect(url_for('workload_settings'))
                    
                success_count = 0
                skip_count = 0
                
                for row in rows[1:]:
                    try:
                        if len(row) <= max(col_indices.values()):
                            skip_count += 1
                            continue
                            
                        item_name = row[col_indices['item_name']]
                        score_val = row[col_indices['score']]
                        
                        if not item_name or score_val is None:
                            skip_count += 1
                            continue
                            
                        try:
                            score = float(score_val)
                        except ValueError:
                            skip_count += 1
                            continue
                            
                        # Upsert logic
                        existing = WorkloadScore.query.filter_by(item_name=item_name).first()
                        if existing:
                            existing.score = score
                        else:
                            new_score = WorkloadScore(item_name=item_name, score=score)
                            db.session.add(new_score)
                            
                        success_count += 1
                        
                    except Exception:
                        skip_count += 1
                        continue
                        
                db.session.commit()
                flash(f'导入成功 {success_count} 条，跳过 {skip_count} 条')
                
            except Exception as e:
                flash(f'导入失败: {str(e)}')
        else:
            flash('不支持的文件格式')
            
        return redirect(url_for('workload_settings'))

    @app.route('/workload/export_scores')
    @login_required
    def export_workload_scores():
        if current_user.role not in ['admin', 'super_admin']:
            flash('无权操作')
            return redirect(url_for('workload_settings'))
            
        export_type = request.args.get('type', 'all') # 'all' or 'unconfigured'
        
        # Logic similar to workload_settings to get unified list
        scores = WorkloadScore.query.all()
        score_map = {s.item_name: s for s in scores}
        
        record_items = db.session.query(WorkloadRecord.item_name).distinct().all()
        record_item_names = [r[0] for r in record_items]
        
        all_names = set(score_map.keys()) | set(record_item_names)
        
        unified_list = []
        for name in all_names:
            score_obj = score_map.get(name)
            is_configured = score_obj is not None
            
            if export_type == 'unconfigured' and is_configured:
                continue
                
            unified_list.append({
                'item_name': name,
                'score': score_obj.score if score_obj else None,
                'is_configured': is_configured
            })
            
        # Sort: Unconfigured first, then Name
        unified_list.sort(key=lambda x: (x['is_configured'], x['item_name']))
        
        # Create Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "分值设定"
        
        # Header
        ws.append(['检查项目', '分值', '状态'])
        
        for item in unified_list:
            ws.append([
                item['item_name'],
                item['score'] if item['score'] is not None else '',
                '已配置' if item['is_configured'] else '未配置'
            ])
            
        # Adjust column width
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        
        # Format: 检查项目与分值对应关系_YYYYMMDDHHMM.xlsx
        timestamp = datetime.now().strftime('%Y%m%d%H%M')
        filename = f"检查项目与分值对应关系_{timestamp}.xlsx"
        
        from flask import make_response
        
        # Use make_response with bytes for full control
        response = make_response(out.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        # RFC 5987 with ASCII fallback
        # Ensure filename is properly quoted
        filename_utf8 = quote(filename)
        response.headers['Content-Disposition'] = f"attachment; filename=\"workload_scores.xlsx\"; filename*=UTF-8''{filename_utf8}"
        
        # Disable caching
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
            
        return response

    @app.route('/workload/settings', methods=['GET', 'POST'])
    @login_required
    def workload_settings():
        if current_user.role not in ['admin', 'super_admin']:
            flash('无权访问')
            return redirect(url_for('workload_stats'))
            
        if request.method == 'POST':
            item_name = request.form.get('item_name')
            score = request.form.get('score')
            
            if item_name and score:
                existing = WorkloadScore.query.filter_by(item_name=item_name).first()
                if existing:
                    existing.score = float(score)
                    flash('分值已更新')
                else:
                    new_score = WorkloadScore(item_name=item_name, score=float(score))
                    db.session.add(new_score)
                    flash('分值已添加')
                db.session.commit()
            return redirect(url_for('workload_settings'))

        scores = WorkloadScore.query.all()
        score_map = {s.item_name: s for s in scores}
        
        # Get all unique items from records
        record_items = db.session.query(WorkloadRecord.item_name).distinct().all()
        record_item_names = [r[0] for r in record_items]
        
        # Merge all unique names
        all_names = set(score_map.keys()) | set(record_item_names)
        
        # Create unified list
        unified_list = []
        for name in all_names:
            score_obj = score_map.get(name)
            unified_list.append({
                'item_name': name,
                'score': score_obj.score if score_obj else None,
                'is_configured': score_obj is not None
            })
            
        # Sort: By Name using Pinyin
        unified_list.sort(key=lambda x: ''.join(lazy_pinyin(x['item_name'])))

        return render_template('workload_settings.html', 
                             items=unified_list, 
                             title='分值设定')

    @app.route('/workload/export_stats')
    @login_required
    def export_workload_stats():
        import calendar
        from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
        from app.models.holidays import Holiday
        
        year, month = get_date_params()
        month_str = f"{year}-{month:02d}"
        selected_month_start = date(year, month, 1)

        # Re-fetch doctors logic
        doctors = Doctor.query.filter(
            and_(
                Doctor.title == '医师',
                or_(
                    Doctor.status == 'active',
                    and_(
                        Doctor.status != 'active',
                        Doctor.resignation_date >= selected_month_start
                    )
                )
            )
        ).order_by(Doctor.display_order.asc()).all()
        
        # Get number of days in month
        _, num_days = calendar.monthrange(year, month)
        month_end = date(year, month, num_days)
        
        # Fetch Holidays
        holidays = Holiday.query.filter(
            Holiday.date >= selected_month_start,
            Holiday.date <= month_end
        ).all()
        holiday_map = {h.date.day: h.type for h in holidays}
        
        # Create Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{year}年{month}月工作量"
        
        # Styles
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')
        header_font = Font(bold=True)
        
        # --- Headers ---
        # Row 1: Dates
        # Row 2: Weekdays
        
        # Fixed Columns: Index, Name
        ws.merge_cells('A1:A2')
        ws['A1'] = '序号'
        ws.merge_cells('B1:B2')
        ws['B1'] = '姓名'
        
        weekday_map = {0: '周一', 1: '周二', 2: '周三', 3: '周四', 4: '周五', 5: '周六', 6: '周日'}
        
        # Daily Columns
        for day in range(1, num_days + 1):
            col_idx = 2 + day # A=1, B=2, so 1st day is C(3)
            cell = ws.cell(row=1, column=col_idx, value=day)
            
            # Weekday
            dt = date(year, month, day)
            wd = weekday_map[dt.weekday()]
            ws.cell(row=2, column=col_idx, value=wd)
            
            # Color Logic:
            # 1. Holiday record exists? -> 'holiday' = Red, 'workday' = Black
            # 2. Weekend? -> Red
            # 3. Default -> Black
            
            is_red = False
            if day in holiday_map:
                if holiday_map[day] == 'holiday':
                    is_red = True
                else: # 'workday' (makeup)
                    is_red = False
            else:
                if dt.weekday() >= 5: # Weekend
                    is_red = True
                    
            if is_red:
                ws.cell(row=1, column=col_idx).font = Font(color="DC3545", bold=True)
                ws.cell(row=2, column=col_idx).font = Font(color="DC3545", bold=True)
                
        # Total Column
        total_col_idx = 2 + num_days + 1
        ws.merge_cells(start_row=1, start_column=total_col_idx, end_row=2, end_column=total_col_idx)
        ws.cell(row=1, column=total_col_idx, value='总计')
        
        # --- Data ---
        # Fetch all records for the month
        records = db.session.query(WorkloadRecord.doctor_id, WorkloadRecord.report_date, WorkloadScore.score)\
            .join(WorkloadScore, WorkloadRecord.item_name == WorkloadScore.item_name)\
            .filter(func.strftime('%Y-%m', WorkloadRecord.report_date) == month_str)\
            .all()
            
        # Aggregate: doctor_id -> day -> total_score
        data_map = {}
        for doc_id, r_date, score in records:
            if doc_id not in data_map:
                data_map[doc_id] = {}
            day = r_date.day
            data_map[doc_id][day] = data_map[doc_id].get(day, 0) + score

        # Write Rows
        for idx, doctor in enumerate(doctors, 1):
            row_idx = 2 + idx
            
            # Index & Name
            ws.cell(row=row_idx, column=1, value=idx)
            ws.cell(row=row_idx, column=2, value=doctor.name)
            
            doc_total = 0
            doc_data = data_map.get(doctor.id, {})
            
            # Daily Scores
            for day in range(1, num_days + 1):
                score = doc_data.get(day, 0)
                col_idx = 2 + day
                # Show 0 if no score, or empty? Screenshot shows 0.
                # But let's show 0 for consistency with screenshot.
                ws.cell(row=row_idx, column=col_idx, value=score if score > 0 else 0)
                doc_total += score
                
            # Total
            ws.cell(row=row_idx, column=total_col_idx, value=round(doc_total, 2))

        # --- Formatting ---
        # Apply borders and alignment to all cells
        max_row = 2 + len(doctors)
        max_col = total_col_idx
        
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = thin_border
                cell.alignment = center_align
                if cell.row <= 2:
                    # Re-apply font logic to ensure it persists and isn't overwritten by default
                    # Check if it's a weekend/holiday column
                    is_weekend_col = False
                    if cell.column > 2 and cell.column <= (2 + num_days):
                        day = cell.column - 2
                        dt = date(year, month, day)
                        
                        # Same logic as above
                        if day in holiday_map:
                            if holiday_map[day] == 'holiday':
                                is_weekend_col = True
                        else:
                            if dt.weekday() >= 5:
                                is_weekend_col = True
                            
                    if is_weekend_col:
                        cell.font = Font(color="DC3545", bold=True)
                    else:
                        cell.font = Font(bold=True)

        # Column Widths
        ws.column_dimensions['A'].width = 6  # Index
        ws.column_dimensions['B'].width = 12 # Name
        for day in range(1, num_days + 1):
            col_letter = openpyxl.utils.get_column_letter(2 + day)
            ws.column_dimensions[col_letter].width = 5 # Narrow for days
        
        total_letter = openpyxl.utils.get_column_letter(total_col_idx)
        ws.column_dimensions[total_letter].width = 10

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d%H%M')
        filename = f"{year}年{month}月工作量统计_{timestamp}.xlsx"
        
        from flask import make_response
        response = make_response(out.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        filename_utf8 = quote(filename)
        response.headers['Content-Disposition'] = f"attachment; filename=\"workload_stats.xlsx\"; filename*=UTF-8''{filename_utf8}"
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        
        return response

    @app.route('/workload/delete_monthly', methods=['POST'])
    @login_required
    def delete_monthly_workload():
        if current_user.role not in ['admin', 'super_admin']:
            return {'success': False, 'message': '无权操作'}, 403
            
        data = request.get_json()
        year = data.get('year')
        month = data.get('month')
        
        if not year or not month:
            return {'success': False, 'message': '参数错误'}, 400
            
        try:
            month_str = f"{year}-{int(month):02d}"
            
            # Delete records for the specified month
            count = WorkloadRecord.query.filter(
                func.strftime('%Y-%m', WorkloadRecord.report_date) == month_str
            ).delete(synchronize_session=False)
            
            db.session.commit()
            return {'success': True, 'message': f'成功删除 {count} 条记录'}
            
        except Exception as e:
            db.session.rollback()
            return {'success': False, 'message': str(e)}, 500

    @app.route('/workload/stats/<int:doctor_id>/details')
    @login_required
    def get_doctor_workload_details(doctor_id):
        year, month = get_date_params()
        month_str = f"{year}-{month:02d}"
        
        doctor = Doctor.query.get_or_404(doctor_id)
        
        # Query records with scores
        records = db.session.query(WorkloadRecord.item_name, WorkloadScore.score)\
            .join(WorkloadScore, WorkloadRecord.item_name == WorkloadScore.item_name)\
            .filter(WorkloadRecord.doctor_id == doctor.id)\
            .filter(func.strftime('%Y-%m', WorkloadRecord.report_date) == month_str)\
            .all()
            
        # Aggregate
        stats = {}
        for item_name, score in records:
            if item_name not in stats:
                stats[item_name] = {'count': 0, 'score': score, 'total': 0}
            stats[item_name]['count'] += 1
            stats[item_name]['total'] += score
            
        # Convert to list
        details = []
        for name, data in stats.items():
            details.append({
                'name': name,
                'count': data['count'],
                'score': data['score'],
                'total': round(data['total'], 2)
            })
            
        # Sort by total score desc
        details.sort(key=lambda x: x['total'], reverse=True)
        
        from flask import jsonify
        return jsonify({
            'doctor_name': doctor.name,
            'period': f"{year}年{month}月",
            'items': details
        })

    @app.route('/workload/doctor/<int:doctor_id>/calendar')
    @login_required
    def doctor_workload_calendar(doctor_id):
        from app.models.holidays import Holiday
        
        year, month = get_date_params()
        month_str = f"{year}-{month:02d}"
        
        doctor = Doctor.query.get_or_404(doctor_id)
        
        # Query all records for this doctor in this month
        records = db.session.query(WorkloadRecord, WorkloadScore.score)\
            .outerjoin(WorkloadScore, WorkloadRecord.item_name == WorkloadScore.item_name)\
            .filter(WorkloadRecord.doctor_id == doctor.id)\
            .filter(func.strftime('%Y-%m', WorkloadRecord.report_date) == month_str)\
            .all()
            
        # Fetch Holidays for this month
        month_start = date(year, month, 1)
        import calendar
        _, num_days = calendar.monthrange(year, month)
        month_end = date(year, month, num_days)
        
        holidays = Holiday.query.filter(
            Holiday.date >= month_start,
            Holiday.date <= month_end
        ).all()
        
        # Map: day -> {type, name}
        holiday_map = {h.date.day: {'type': h.type, 'name': h.name} for h in holidays}
            
        # Aggregate by day
        daily_stats = {}
        total_items = 0
        total_score = 0
        
        # Initialize all days
        for day in range(1, num_days + 1):
            daily_stats[day] = {'item_count': 0, 'score': 0}
            
        for record, score in records:
            day = record.report_date.day
            daily_stats[day]['item_count'] += 1
            s = score if score is not None else 0
            daily_stats[day]['score'] += s
            
            total_items += 1
            total_score += s
            
        # Format for template
        calendar_data = []
        # Calculate start day of week (0=Mon, 6=Sun)
        start_day_weekday = date(year, month, 1).weekday()
        
        # Add empty slots for days before start of month
        for _ in range(start_day_weekday):
            calendar_data.append(None)
            
        for day in range(1, num_days + 1):
            # Determine status: 'workday', 'weekend', 'holiday'
            # Default
            current_date = date(year, month, day)
            is_weekend = current_date.weekday() >= 5 # 5=Sat, 6=Sun
            
            status = 'workday'
            if is_weekend:
                status = 'weekend'
                
            holiday_name = None
            
            # Override with holiday data
            if day in holiday_map:
                h_data = holiday_map[day]
                h_type = h_data['type']
                if h_type == 'holiday':
                    status = 'holiday'
                    holiday_name = h_data['name']
                elif h_type == 'workday': # Makeup day
                    status = 'workday'
                    holiday_name = h_data['name'] # Might be "班" or specific name
            
            calendar_data.append({
                'day': day,
                'item_count': daily_stats[day]['item_count'],
                'score': round(daily_stats[day]['score'], 2),
                'status': status,
                'holiday_name': holiday_name
            })
            

            
        # Ensure complete weeks (multiple of 7)
        while len(calendar_data) % 7 != 0:
            calendar_data.append(None)
            
        # Round total score
        total_score = round(total_score, 2)
        
        return render_template('doctor_workload_calendar.html',
                             doctor=doctor,
                             current_year=year,
                             current_month=month,
                             calendar_data=calendar_data,
                             total_items=total_items,
                             total_score=total_score,
                             title=f'{doctor.name} - 工作量日历')

    @app.route('/workload/doctor/<int:doctor_id>/daily/<int:year>/<int:month>/<int:day>')
    @login_required
    def doctor_daily_workload(doctor_id, year, month, day):
        doctor = Doctor.query.get_or_404(doctor_id)
        target_date = date(year, month, day)
        
        # Query records for this specific day
        records = db.session.query(WorkloadRecord.item_name, WorkloadScore.score)\
            .outerjoin(WorkloadScore, WorkloadRecord.item_name == WorkloadScore.item_name)\
            .filter(WorkloadRecord.doctor_id == doctor.id)\
            .filter(func.date(WorkloadRecord.report_date) == target_date)\
            .all()
            
        # Aggregate
        stats = {}
        for item_name, score in records:
            s = score if score is not None else 0
            if item_name not in stats:
                stats[item_name] = {'count': 0, 'score': s, 'total': 0}
            stats[item_name]['count'] += 1
            stats[item_name]['total'] += s
            
        # Convert to list
        daily_items = []
        total_items = 0
        total_score = 0
        
        for name, data in stats.items():
            daily_items.append({
                'name': name,
                'count': data['count'],
                'score': data['score'],
                'total': round(data['total'], 2)
            })
            total_items += data['count']
            total_score += data['total']
            
        # Sort by: Count (Desc), Score (Desc), Name (Asc)
        daily_items.sort(key=lambda x: (-x['count'], -x['score'], x['name']))
            
        total_score = round(total_score, 2)
        
        return render_template('doctor_daily_workload.html',
                             doctor=doctor,
                             date=target_date,
                             daily_items=daily_items,
                             total_items=total_items,
                             total_score=total_score,
                             title=f'{doctor.name} - {year}年{month}月{day}日工作量')

    @app.route('/workload/doctor/<int:doctor_id>/monthly/<int:year>/<int:month>')
    @login_required
    def doctor_monthly_details(doctor_id, year, month):
        doctor = Doctor.query.get_or_404(doctor_id)
        month_str = f"{year}-{month:02d}"
        
        # Query records with scores for the month
        records = db.session.query(WorkloadRecord.item_name, WorkloadScore.score)\
            .join(WorkloadScore, WorkloadRecord.item_name == WorkloadScore.item_name)\
            .filter(WorkloadRecord.doctor_id == doctor.id)\
            .filter(func.strftime('%Y-%m', WorkloadRecord.report_date) == month_str)\
            .all()
            
        # Aggregate
        stats = {}
        for item_name, score in records:
            if item_name not in stats:
                stats[item_name] = {'count': 0, 'score': score, 'total': 0}
            stats[item_name]['count'] += 1
            stats[item_name]['total'] += score
            
        # Convert to list
        details = []
        grand_total_items = 0
        grand_total_score = 0
        
        for name, data in stats.items():
            details.append({
                'name': name,
                'count': data['count'],
                'score': data['score'],
                'total': round(data['total'], 2)
            })
            grand_total_items += data['count']
            grand_total_score += data['total']
            
        # Sort by: Count (Desc), Score (Desc), Name (Asc)
        details.sort(key=lambda x: (-x['count'], -x['score'], x['name']))
        
        grand_total_score = round(grand_total_score, 2)
        
        return render_template('doctor_monthly_details.html',
                             doctor=doctor,
                             year=year,
                             month=month,
                             details=details,
                             total_items=grand_total_items,
                             total_score=grand_total_score,
                             title=f'{doctor.name} - {year}年{month}月工作量详情')
